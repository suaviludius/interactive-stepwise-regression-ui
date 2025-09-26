#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import os # Модуль для работы с файловой системой (функции операционной системы)
from datetime import datetime
import random
import numpy as np

import openpyxl
from openpyxl.styles import PatternFill

# -- GUI -----------------------------------------------------------
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui     import QIcon
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QPushButton, QTextEdit

# -- Import Regression & Design -------------------------------------------------
# Добавляем родительскую папку в путь
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.core import StepwiseRegressionEngine   # Class Regression
from src.ui import FileHandler, FramelessWindow, Ui_MainWindow, StyleSheetTableDatabase, StyleSheetStatusBar, StyleSheetScrollBar, StyleSheetApp

# -- Import matplot ------------------------------------------------
import matplotlib
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt

class MainWindow(QMainWindow):
    ###################################
    # [ Инициализация рабочего окна ] #
    ###################################
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)  # Установка GUI

        self.setSB() # Инициализируем statusBar
        self.actionsDWTBMB()  # Обработка действий dockWidget, toolBar, menuBar
        self.actionsMW()  # Обработка действий mainWindow
        self.graphicsInit() # Инициализация графика

        self.ui.dockExplorer.setVisible(False)
        self.ui.dockAnalyse.setVisible(False)
        self.ui.dockGraphics.setVisible(False)

        self.show() # Отображение того, что получилось

    ####################################
    # [ Инициализаторы элементов окна] #
    ####################################

    # Строка состояния
    def setSB(self):
        # Настройки для statusBar
        self.statusBar = QStatusBar()
        self.b = QPushButton("click here")
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("Смотри на меня")
        self.statusBar.setStyleSheet(StyleSheetStatusBar)

    # Файловое дерево для "ПРОВОДНИК"
    def setFT(self,path):
        self.model = QtWidgets.QFileSystemModel()
        self.model.setRootPath(QtCore.QDir.rootPath())
        self.model.setNameFilters(["*.xlsx","*.tree","*.rep"])  # Создание фильтра на типы отображаемых файлов
        self.model.setNameFilterDisables(False) # Сокрытие типов файла не входящих в фильтр

        self.FT = FileHandler(self, self.ui.treeView, self.model, self.ui.actionAnalyseTB, self.ui.actionReportTB);
        self.FT.setFT(path)

    #####################################################
    # [ Обработка кнопок dockWidget, toolBar, menuBar ] #
    #####################################################

    def actionsDWTBMB(self):
        # Группируем действия левого toolBar, чтобы в момент времени было активно только 1 действие
        groupTB =  QActionGroup(self)
        groupTB.addAction(self.ui.actionHomeTB)
        groupTB.addAction(self.ui.actionAnalyseTB)
        groupTB.addAction(self.ui.actionReportTB)
        self.ui.actionHomeTB.setChecked(True)

        # Обработка нажатия кнопки "Х" для dockWidget
        self.ui.dockExplorer.visibilityChanged.connect(lambda: self.ui.actionExplorerMB.setChecked(self.ui.dockExplorer.isVisible()))
        self.ui.dockAnalyse.visibilityChanged.connect(lambda: self.ui.actionAnalyseMB.setChecked(self.ui.dockAnalyse.isVisible()))
        self.ui.dockGraphics.visibilityChanged.connect(lambda: self.ui.actionGraphicsMB.setChecked(self.ui.dockGraphics.isVisible()))

        # Обработка нажатия кнопок для левого toolBar
        self.ui.actionHomeTB.triggered.connect(lambda: self.ui.mainCenterBodyStacked.setCurrentWidget(self.ui.pageHome))
        self.ui.actionAnalyseTB.triggered.connect(lambda: self.ui.mainCenterBodyStacked.setCurrentWidget(self.ui.pageAnalyse))
        self.ui.actionReportTB.triggered.connect(lambda: self.ui.mainCenterBodyStacked.setCurrentWidget(self.ui.pageReport))

        # Обработка нажатия кнопок для верхнего menuBar
        self.ui.actionOpenFileMB.triggered.connect(lambda: self.selectFile())
        self.ui.actionOpenFolderMB.triggered.connect(lambda: self.selectFolder())
        self.ui.actionSave.triggered.connect(lambda: self.saveFile())
        self.ui.actionExplorerMB.triggered.connect(lambda:self.ui.dockExplorer.setVisible(self.ui.actionExplorerMB.isChecked()))
        self.ui.actionAnalyseMB.triggered.connect(lambda: self.ui.dockAnalyse.setVisible(self.ui.actionAnalyseMB.isChecked()))
        self.ui.actionGraphicsMB.triggered.connect(lambda: self.ui.dockGraphics.setVisible(self.ui.actionGraphicsMB.isChecked()))
        self.ui.actionInstructionMB.triggered.connect(self.showInstruction)

    def saveFile(self):
        """
        Сохраняет Excel файл с цветными столбцами с проверкой существования файла
        """
        # Проверяем существование файла
        try:
        # Загружаем workbook
            wb = openpyxl.load_workbook(self.MR.FileName)
            ws = wb.active

            # Создаем стили заливки
            green_fill = PatternFill(start_color='bcf5d3', end_color='bcf5d3', fill_type='solid')
            red_fill = PatternFill(start_color='ffc5c5', end_color='ffc5c5', fill_type='solid')

            print(self.MR.IndX_ADD, self.MR.IndX_DEL)
            # Применяем цвета к столбцам
            for col_idx in self.MR.IndX_ADD:
                col_letter = openpyxl.utils.get_column_letter(col_idx+1)
                for cell in ws[col_letter]:
                    cell.fill = green_fill

            for col_idx in self.MR.IndX_DEL:
                col_letter = openpyxl.utils.get_column_letter(col_idx+1)
                for cell in ws[col_letter]:
                    cell.fill = red_fill

            # Диалог выбора места сохранения
            options = QFileDialog.Options()
            default_filename = f"colored_{os.path.basename(self.MR.FileName)}"

            output_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить цветной Excel файл",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)",
                options=options
            )

            # Если пользователь нажал "Отмена"
            if not output_path:
                self.statusBar.showMessage("Сохранение отменено")
                return None

            # Добавляем расширение .xlsx если его нет
            if not output_path.lower().endswith('.xlsx'):
                output_path += '.xlsx'

            # Сохраняем файл
            wb.save(output_path)
            self.statusBar.showMessage(f"Файл сохранен: {output_path}")
            return output_path

        except Exception as e:
            self.statusBar.showMessage(f"Ошибка при сохранении файла: {str(e)}")
            return None

    def showInstruction(self):
        """Показывает инструкцию по использованию приложения."""
        instruction_text = """
    <h1>Инструкция по использованию приложения</h1>
    <h2>Анализ множественной линейной регрессии</h2>

    <h3> 1. Загрузка данных</h3>
    <p><b>Шаг 1:</b> Нажмите кнопку "Файл" на главной странице или перейдите в меню "Файл" и нажмите "Выбрать файл" или "Выбрать папку"</p>
    <p><b>Шаг 2:</b> Выберите файл данных в формате .xlsx</p>
    <p><b>Требования к данным:</b></p>
    <ul>
        <li>Лист с данными должен называться "Лист1"</li>
        <li>Независимые переменные (X) должны располагаться в первых столбцах</li>
        <li>Зависимые переменные (Y) - в последних столбцах</li>
        <li>Первая строка должна содержать названия переменных</li>
    </ul>

    <h3> 2. Настройка анализа</h3>
    <p><b>Шаг 3:</b> Перейдите в раздел "Анализ" (вторая кнопка в левой панели)</p>
    <p><b>Шаг 4:</b> В виджете "Анализ" выполните:</p>
    <ul>
        <li>Выберите количество зависимых переменных</li>
        <li>Выберите конкретную зависимую переменную для анализа</li>
        <li>Выберите метод отбора переменных</li>
        <li>Выберите режим использования алгоритма при необходимости</li>
    </ul>

    <h3> 3. Выполнение пошагового анализа</h3>
    <p><b>Методы анализа:</b></p>
    <ul>
        <li><b>Обратное исключение</b> - начинается с полной модели, последовательно удаляет наименее значимые переменные</li>
        <li><b>Прямое включение</b> - начинается с пустой модели, последовательно добавляет наиболее значимые переменные</li>
    </ul>
    <p><b>Режимы алгоритма:</b></p>
    <ul>
        <li><b>Авто отбор</b> - выполняет алгоритмы отбора до достижения критериев:
        коэффициент статистики (КС) превышает заданный порог (КС Const), а коэффициент детерминации (КД)
        не ниже минимального значения (КД Const)</li>
        <li><b>Пошаговый отбор</b> - пошаговое выполнение метода без ограничений</li>
    </ul>

    <p><b>Управление процессом:</b></p>
    <ul>
        <li>Кнопка <b>"Вперед"</b> - выполнить следующий шаг алгоритма</li>
        <li>Кнопка <b>"Назад"</b> - отменить последний шаг</li>
        <li>Кнопка <b>"Заполнить"</b> - включить все переменные в модель</li>
        <li>Кнопка <b>"Очистить"</b> - удалить все переменные из модели</li>
        <li>Кнопка <b>"+"</b> - добавить независимую переменную</li>
        <li>Кнопка <b>"-"</b> - удалить независимую переменную</li>
    </ul>

    <h3> 4. Анализ результатов</h3>
    <p><b>В реальном времени отслеживайте:</b></p>
    <ul>
        <li>Коэффициент детерминации (КД или R²)</li>
        <li>Коэффициент статистики (КС или F-статистику)ы </li>
        <li>Стандартную ошибку оценки (ОД, НОД, ОСКО)</li>
        <li>Графики изменения метрик</li>
    </ul>

    <h3> 5. Сохранение результатов</h3>
    <p><b>Шаг 5:</b> Перейдите в раздел "Отчет"</p>
    <p><b>Шаг 6:</b> Нажмите "Создать отчет" для сохранения результатов в файл .rep</p>

    <h3> Советы по использованию</h3>
    <ul>
        <li>Используйте автоматический режим для быстрого анализа, по заранее выставленным порогам (КС и КД Const)</li>
        <li>Ручной режим позволяет контролировать каждый шаг</li>
        <li>Следите за изменением R² - он не должен значительно уменьшаться</li>
        <li>Анализируйте графики для визуальной оценки процесса</li>
    </ul>

    <h3> Возможные проблемы и решения</h3>
    <ul>
        <li><b>Файл не загружается</b> - проверьте формат и структуру данных</li>
        <li><b>Алгоритм не работает</b> - убедитесь, что выбрана зависимая переменная</li>
        <li><b>Некорректные результаты</b> - проверьте мультиколлинеарность переменных</li>
    </ul>

    <p style="color: #3c90a4; font-weight: bold;"> Для дополнительной помощи: denisDanilkopatich@gmail.com</p>
    """

        # Создаем диалоговое окно с инструкцией
        dialog = QDialog(self)
        dialog.setWindowTitle("Инструкция по использованию приложения")
        dialog.setMinimumSize(800, 600)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
                color: #ffffff;
            }
            QScrollArea {
                border: none;
            }
        """)

        # Создаем текстовое поле с поддержкой HTML
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setHtml(instruction_text)
        text_edit.setStyleSheet("""
            QTextEdit {
                background-color: #2b2b2b;
                color: #ffffff;
                border: none;
                font-family: Arial, sans-serif;
                font-size: 12px;
                padding: 10px;
            }
        """)

        # Кнопка закрытия
        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(dialog.accept)
        close_button.setStyleSheet("""
            QPushButton {
                background-color: #3c90a4;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #255a66;
            }
        """)

        # Компоновка
        layout = QVBoxLayout()
        layout.addWidget(text_edit)
        layout.addWidget(close_button)
        dialog.setLayout(layout)

        # Показываем диалог
        dialog.exec_()

    #############################################
    # [ Обработка кнопок dockWidget - Analyse ] #
    #############################################

    def actionsDWA(self):
        # Группируем действия dockWidget Auto-Hand, чтобы в момент времени было активно только 1 действие
        groupModeAH =  QButtonGroup(self)
        groupModeAH.addButton(self.ui.buttonAutoMode)
        groupModeAH.addButton(self.ui.buttonHandMode)
        self.ui.buttonAutoMode.setChecked(True)

        # Группируем действия dockWidget Back-Step, чтобы в момент времени было активно только 1 действие
        groupModeBS =  QButtonGroup(self)
        groupModeBS.addButton(self.ui.buttonBackEleminate)
        groupModeBS.addButton(self.ui.buttonStepwise)
        self.ui.buttonBackEleminate.setChecked(True)

        # Кнопки сворачивания пмодулей dockWidget
        self.ui.pushButtonActions.clicked.connect(lambda: self.ui.actions.setVisible(self.ui.pushButtonActions.isChecked()))
        self.ui.pushButtonModes.clicked.connect(lambda: self.ui.modes.setVisible(self.ui.pushButtonModes.isChecked()))
        self.ui.pushButtonConstants.clicked.connect(lambda: self.ui.constParameters.setVisible(self.ui.pushButtonConstants.isChecked()))
        self.ui.pushButtonParameters.clicked.connect(lambda: self.ui.parameters.setVisible(self.ui.pushButtonParameters.isChecked()))

        self.ui.countYComboBox.currentIndexChanged.connect(self.setCountAndChoiceDepVar) # Выбор числа зависимых переменных
        self.ui.choiceYComboBox.currentIndexChanged.connect(self.setChoiceDepVar)   # Выбор зависимой переменной
        self.ui.buttonClean.clicked.connect(self.setCleanTable) # Заполнение таблицы
        self.ui.buttonFill.clicked.connect(self.setFillTable)   # Очищение таблицы

        # Основные кнопки выполнения регрессии
        self.ui.buttonBack.clicked.connect(self.setBackup)  # ШАГ: назад
        self.ui.buttonForward.clicked.connect(self.setAnalyseStep)  # ШАГ: вперед

    #------------------------------------------#

    # Сворачивание меню
    def toggleFrameAnalyse(self, frame):
        if self.ui.dockExplorer.setVisible(self.ui.actionExplorerMB.isChecked()):
            frame.hide()
            self.is_frame_visible = False
            self.button.setIcon(QIcon("arrow_icon_up.png"))
        else:
            frame.show()
            self.is_frame_visible = True
            self.button.setIcon(QIcon("arrow_icon_down.png"))

    # Выбор зависимой переменной
    def setChoiceDepVar(self):
        self.MR.Ynum = self.ui.choiceYComboBox.currentIndex()                                                                 # Номер столбца Y для объекта MR
        self.Ynum = self.MR.Columns - self.ui.countYComboBox.currentIndex() + self.ui.choiceYComboBox.currentIndex() - 1    # Номер столбца Y для функции закрашивания
        self.changeBackgrounColor(list(range(self.MR.ColumnsX_BUF, self.MR.Columns)),'#505A6E','#f6f9ff')
        self.changeBackgrounColor([self.Ynum],'#255a66','#f6f9ff')
        self.setFillTable()

    # Выбор количества зависимых переменных
    def setCountDepVar(self):
        self.MR.readExel(self.ui.countYComboBox.currentIndex()+1)
        self.ui.choiceYComboBox.clear()
        self.ui.choiceYComboBox.addItems(self.MR.dataset.columns[self.MR.ColumnsX:self.MR.Columns])       # Ассортимент зависимых переменных
        self.setFillTable()
        self.changeBackgrounColor(list(range(self.MR.ColumnsX_BUF, self.MR.Columns)),'#505A6E','#f6f9ff')

    # Выбор количества зависимых переменных и самой зависимой переменной
    def setCountAndChoiceDepVar(self):
        self.setCountDepVar()
        self.setChoiceDepVar()

    # Отчистить содержимое таблицы (красный цвет - элемент удаоен)
    def setCleanTable(self):
        self.MR.cleanTable()
        self.setEditTextBox()
        self.changeBackgrounColor(list(range(self.MR.ColumnsX_BUF)),'#ffc5c5','#3e4556')
        self.graphicsDraw()

    # Заполнить таблицу данными (зеленый цвет - элемент добавлен)
    def setFillTable(self):
        self.MR.filledTable()
        self.setEditTextBox()
        self.changeBackgrounColor(list(range(self.MR.ColumnsX_BUF)),'#bcf5d3','#3e4556')
        self.graphicsDraw()

    # Кнопка возврата действия
    def setBackup(self):
        self.MR.outBackup() 
        for i in range(len(self.MR.X_BUF[0])):
            if i in self.MR.IndX_ADD: self.changeBackgrounColor([i],'#bcf5d3','#3e4556')
            else: self.changeBackgrounColor([i],'#ffc5c5','#3e4556')
        self.setEditTextBox()
        self.graphicsDraw()

    # Основная кнопка действия регрессии
    def setAnalyseStep(self):
        if(self.ui.buttonBackEleminate.isChecked()):
            if(self.ui.buttonAutoMode.isChecked()):
                self.backwardElimination()
            elif(self.ui.buttonHandMode.isChecked()):
                self.backwardEliminationStep()
        elif(self.ui.buttonStepwise.isChecked()):
            if(self.ui.buttonAutoMode.isChecked()):
                self.stepwise()
            elif(self.ui.buttonHandMode.isChecked()):
                self.stepwiseStep()
        self.graphicsDraw()

    #------------------------------------------#

    # Цикл братной ликвидации
    def backwardElimination(self):
        self.statusBar.showMessage("Обратная ликвидация")
        while self.MR.R2 > self.MR.Const_R2 and len(self.MR.IndX_ADD) > 1:
            self.backwardEliminationStep()

    # Шаг обратной ликвидации
    def backwardEliminationStep(self):
        if len(self.MR.IndX_ADD) > 1:
            self.MR.DELXK()
            self.setEditTextBox()
            self.changeBackgrounColor([self.MR.IndX_DEL[-1]],'#ffc5c5','#3e4556')
            self.statusBar.showMessage(f'Был удален элемент: {self.MR.IndX_DEL[-1] + 1}')

    # Цикл прямого включения
    def stepwise(self):
        self.statusBar.showMessage("Прямое включение")
        while self.MR.FSKF < self.MR.Const_FSKF and len(self.MR.IndX_DEL) > 1:
            self.stepwiseStep()

    # Шаг прямого включения
    def stepwiseStep(self):
        self.MR.ADDXK()
        self.setEditTextBox()
        self.changeBackgrounColor([self.MR.IndX_ADD[-1]],'#bcf5d3','#3e4556')
        self.statusBar.showMessage(f'Был добавлен элемент: {self.MR.IndX_ADD[-1] + 1}')

    # Включения фактора по выбору пользователя
    def userStepwiseStep(self):
        i = self.tableDatabaseSelectionItem()
        if i in self.MR.IndX_DEL:
            self.MR.ADDXE(self.MR.IndX_DEL.index(i))
            self.setEditTextBox()
            self.changeBackgrounColor([self.MR.IndX_ADD[-1]],'#bcf5d3','#3e4556')
            self.statusBar.showMessage(f'Был добавлен элемент: {self.MR.IndX_ADD[-1] + 1}')
            self.tableDatabaseSelectionItem()

    # Исключение фактора по выбору пользователя
    def userBackwardStep(self):
        i = self.tableDatabaseSelectionItem()
        if i in self.MR.IndX_ADD:
            self.MR.DELXE(self.MR.IndX_ADD.index(i))
            self.setEditTextBox()
            self.changeBackgrounColor([self.MR.IndX_DEL[-1]],'#ffc5c5','#3e4556')
            self.statusBar.showMessage(f'Был удален элемент: {self.MR.IndX_DEL[-1] + 1}')
            self.tableDatabaseSelectionItem()

    ##############################################
    # [ Обработка кнопок dockWidget - Graphics ] #
    ##############################################

    def actionsDWG(self):
        # Группируем действия dockWidget Auto-Hand, чтобы в момент времени было активно только 1 действие
        groupGraphics =  QButtonGroup(self)
        groupGraphics.addButton(self.ui.buttonGrSKD)
        groupGraphics.addButton(self.ui.buttonGrSKS)
        groupGraphics.addButton(self.ui.buttonGrE)
        self.ui.buttonGrE.setChecked(True)

        self.ui.buttonGrSKD.clicked.connect(self.graphicsDraw) # Значения R2 для обратной ликвидации
        self.ui.buttonGrSKS.clicked.connect(self.graphicsDraw) # Значения R2 для обратной ликвидации
        self.ui.buttonGrE.clicked.connect(self.graphicsDraw) # Ошибки посчитанного и начального Y

    #---------------------------------#

    # Инициализация графического поля
    def graphicsInit(self):
        # -- Graphics init --
        self.ui.figure = plt.figure()
        self.ui.figure.patch.set_facecolor('#3e4556')
        self.ui.canvas = FigureCanvas(self.ui.figure)
        self.ui.verticalLayoutGraphics = QtWidgets.QVBoxLayout(self.ui.Graphics)
        self.ui.verticalLayoutGraphics.setContentsMargins(0, 0, 0, 0)
        self.ui.verticalLayoutGraphics.addWidget(self.ui.canvas)
        matplotlib.rcParams.update({'font.size': 8,'ytick.color':'#E6E6E6', 'xtick.color':'#E6E6E6'})
        plt.rc('axes', facecolor='#3e4556', edgecolor='#E6E6E6', axisbelow=True, grid=True)
        plt.rc('lines', linewidth=2)

    # Отрисовка нужного графика
    def graphicsDraw(self):
        # -- Graphics paint --
        self.ui.figure.clear()
        if(self.ui.buttonGrSKD.isChecked() and len(self.MR.R2_DEL) > 0):
            plt.bar(range(len(self.MR.IndX_DEL)), self.MR.R2_DEL, color ='#3c90a4', width = 0.2, edgecolor='#E6E6E6')
        elif(self.ui.buttonGrSKS.isChecked() and len(self.MR.FSKF_ADD) > 0):
            plt.bar(range(len(self.MR.IndX_ADD)), self.MR.FSKF_ADD, color ='#3c90a4', width = 0.2, edgecolor='#E6E6E6')
        elif(self.ui.buttonGrE.isChecked() and len(self.MR.E) > 0):
            yneg = []
            ypos = []
            for e in self.MR.E:
                if e < 0:
                    yneg.append(e*(-1))
                    ypos.append(0)
                else:
                    yneg.append(0)
                    ypos.append(e)
            y_error = [yneg, ypos] # Ошибки вычислений
            plt.errorbar(x=range(len(self.MR.E)), y=self.MR.Y[:,self.MR.Ynum], yerr=y_error, marker ='o', color ='#3c90a4', ecolor='red', linestyle='none', elinewidth=1)
        self.ui.canvas.draw()


    ###################################
    # [ Обработка кнопок mainWindow ] #
    ###################################

    def actionsMW(self):
        self.ui.buttonSelectFile.clicked.connect(self.selectFile) # Выбор файла с данными
        self.ui.buttonSelectFolder.clicked.connect(self.selectFolder) # Выбор рабочей папки
        self.ui.butonRandomData.clicked.connect(self.selectRandom) # Генерация файла со случайными данными
    #---------------------------------#

    # Выбор рабочего файла
    def selectFile(self):
        pathFile = QFileDialog.getOpenFileName(self,'Выбор файла данных',os.getenv('HOME'),'XLSX Files(*.xlsx)')
        if pathFile[0] == '': return
        pathFolder = os.path.dirname(pathFile[0])
        self.setFT(pathFolder) # Инициализация файлового дерева
        self.statusBar.showMessage("Выбран файл: {}".format(pathFile[0])) # Сообщение в statusBar
        self.ui.buttonSelectFolder.setVisible(False)

        self.FT.openFile(pathFile[0])

    # Выбор рабочей папки
    def selectFolder(self):
        pathFolder = QFileDialog.getExistingDirectory(self,'Выбор рабочей папки',os.getenv('HOME'), QFileDialog.ShowDirsOnly)
        if pathFolder == '': return
        self.setFT(pathFolder) # Инициализация файлового дерева
        self.statusBar.showMessage("Выбрана папка: {}".format(pathFolder)) # Сообщение в statusBar
        self.ui.buttonSelectFolder.setVisible(False)
        self.ui.dockExplorer.setVisible(1)


    def selectRandom(self):
        """
        Создает Excel файл с примером данных X1-X5, Y1-Y3 и случайными числами
        """
        try:
            n_samples = 50  # Увеличиваем количество наблюдений для лучшей регрессии
            n_x = 5
            n_y = 3

            # 1. Генерируем независимые переменные X с контролируемой корреляцией
            np.random.seed(int(datetime.now().timestamp()) % 1000)  # Случайный seed

            # Базовые независимые переменные с разными распределениями
            X = np.zeros((n_samples, n_x))

            # X1: Нормальное распределение (основной фактор)
            X[:, 0] = np.random.normal(50, 15, n_samples)

            # X2: Равномерное распределение (слабая корреляция с X1)
            X[:, 1] = 0.3 * X[:, 0] + np.random.uniform(20, 80, n_samples)

            # X3: Логнормальное распределение (независимый фактор)
            X[:, 2] = np.random.lognormal(3, 0.4, n_samples) * 8

            # X4: Показательное распределение (умеренная корреляция с X1)
            X[:, 3] = 0.5 * X[:, 0] + np.random.exponential(15, n_samples)

            # X5: Гамма распределение (независимый фактор)
            X[:, 4] = np.random.gamma(2, 12, n_samples)

            # Нормализуем и масштабируем X
            X = (X - np.mean(X, axis=0)) / np.std(X, axis=0)  # Стандартизация
            X = 25 + 50 * (X - np.min(X, axis=0)) / (np.max(X, axis=0) - np.min(X, axis=0))  # Масштабирование 0-100
            X = np.clip(X, 1, 99)
            X = np.round(X, 2)

            # 2. Создаем реалистичные коэффициенты регрессии
            # Каждая Y переменная имеет разную зависимость от X
            true_coefficients = np.array([
                # X1   X2   X3   X4   X5   Intercept
                [ 2.5, -1.8,  0.9, -1.2,  0.7,  10.0],  # Y1: сильно зависит от X1, X2
                [-1.2,  1.5, -0.6,  0.8, -0.9,  15.0],  # Y2: сильно зависит от X2, X4
                [ 0.8, -0.7,  1.3, -0.5,  1.1,   8.0]   # Y3: сильно зависит от X3, X5
            ])

            # 3. Генерируем Y переменные на основе линейной модели
            Y = np.zeros((n_samples, n_y))

            # Добавляем столбец единиц для intercept
            X_with_intercept = np.column_stack([np.ones(n_samples), X])

            for j in range(n_y):
                # Линейная комбинация: Y = β0 + β1*X1 + β2*X2 + ... + β5*X5 + ε
                deterministic_part = X_with_intercept @ true_coefficients[j]

                # Добавляем случайную ошибку (10% от диапазона Y)
                noise_std = 0.1 * (np.max(deterministic_part) - np.min(deterministic_part))
                noise = np.random.normal(0, noise_std, n_samples)

                Y[:, j] = deterministic_part + noise

            # Масштабируем Y к реалистичному диапазону
            Y = 5 + 90 * (Y - np.min(Y)) / (np.max(Y) - np.min(Y))
            Y = np.clip(Y, 1, 100)
            Y = np.round(Y, 2)

            # 4. Проверяем пригодность данных для регрессии
            correlation_matrix = np.corrcoef(X.T)
            np.fill_diagonal(correlation_matrix, 0)  # Исключаем диагональ
            max_correlation = np.max(np.abs(correlation_matrix))

            self.statusBar.showMessage(f"Максимальная корреляция между X: {max_correlation:.3f}")

            if max_correlation > 0.8:
                self.statusBar.showMessage("Предупреждение: возможна мультиколлинеарность")

            # Создаем Excel
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet("Лист1", 0)

            headers = [f'X{i+1}' for i in range(n_x)] + [f'Y{i+1}' for i in range(n_y)]
            ws.append(headers)

            for i in range(n_samples):
                row_data = [round(float(X[i, j]), 2) for j in range(n_x)] + \
                        [round(float(Y[i, j]), 2) for j in range(n_y)]
                ws.append(row_data)

            # Генерируем имя файла по умолчанию
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"sample_data_{timestamp}.xlsx"

            # Диалог выбора места сохранения
            options = QFileDialog.Options()
            output_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить пример Excel файла",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)",
                options=options
            )

            # Если пользователь нажал "Отмена"
            if not output_path:
                self.statusBar.showMessage("Создание файла отменено")
                return None

            # Добавляем расширение .xlsx если его нет
            if not output_path.lower().endswith('.xlsx'):
                output_path += '.xlsx'

            # Сохраняем файл
            wb.save(output_path)

            self.statusBar.showMessage(f"Файл создан: {output_path}")
            pathFolder = os.path.dirname(output_path)
            self.setFT(pathFolder) # Инициализация файлового дерева
            self.ui.buttonSelectFolder.setVisible(False)
            self.FT.openFile(output_path)
            return output_path

        except Exception as e:
            self.statusBar.showMessage(f"Ошибка при создании файла: {str(e)}")
            return None

    ###################################
    # [ Обработка кнопок reportWindow ] #
    ###################################

    def actionsRW(self):
        self.ui.reportText.setFont(QFont('MS Shell Dlg 2', 16))  # Установка шрифта размером 16
        self.ui.buttonReportCreate.clicked.connect(self.setReport) # Создание отчета

    #---------------------------------#

    def setReport(self):
        X_DEL = []; X_ADD = [];
        for i in self.MR.IndX_DEL: X_DEL.append(self.MR.dataset.columns[i])
        for i in self.MR.IndX_ADD: X_ADD.append(self.MR.dataset.columns[i])

        file_name, _ = QFileDialog.getSaveFileName(self, 'Сохранить отчет', '', '(*.rep)')

        if file_name:
            try:
                self.saveReport(file_name, X_DEL, X_ADD)

                self.statusBar.showMessage("Отчет успешно сохранен")
                self.openReport(file_name)

            except Exception as e:
                self.statusBar.showMessage(f"Ошибка при сохранении: {str(e)}")

    def saveReport(self, file_name, X_DEL, X_ADD):
        """Сохранение отчета в текстовом формате с красивым форматированием"""
        with open(file_name, 'w', encoding='utf-8') as f:
            f.write("ОТЧЕТ ПО РЕГРЕССИОННОМУ АНАЛИЗУ\n")
            f.write("=" * 50 + "\n")
            f.write(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n\n")

            f.write("ОБЩАЯ СТАТИСТИКА:\n")
            f.write("-" * 25 + "\n")
            f.write(f"Включенные переменные: {len(X_ADD)}\n")
            f.write(f"Исключенные переменные: {len(X_DEL)}\n")
            f.write(f"Коэффициент R²: {self.MR.R2:.6f}\n")
            f.write(f"F-статистика: {self.MR.FSKF:.6f}\n\n")

            f.write("ВКЛЮЧЕННЫЕ ПЕРЕМЕННЫЕ:\n")
            f.write("-" * 25 + "\n")
            for i, var in enumerate(X_ADD, 1):
                f.write(f"{i}. {var}\n")
            if not X_ADD:
                f.write("Нет включенных переменных\n")
            f.write("\n")

            f.write("ИСКЛЮЧЕННЫЕ ПЕРЕМЕННЫЕ:\n")
            f.write("-" * 25 + "\n")
            for i, var in enumerate(X_DEL, 1):
                f.write(f"{i}. {var}\n")
            if not X_DEL:
                f.write("Нет исключенных переменных\n")
            f.write("\n")

            f.write("КОЭФФИЦИЕНТЫ РЕГРЕССИИ:\n")
            f.write("-" * 25 + "\n")
            for i, coef in enumerate(self.MR.REE):
                if i == 0:
                    f.write(f"Коэффициент {i}: {coef:.6f} - Свободный член (константа)\n")
                else:
                    var_index = i - 1  # потому что 0 - константа
                    if var_index < len(X_ADD):
                        var_name = X_ADD[var_index]
                        f.write(f"Коэффициент {i}: {coef:.6f} - Переменная {var_name}\n")
                    else:
                        f.write(f"Коэффициент {i}: {coef:.6f} - Переменная X{var_index}\n")
            f.write("\n")

            # Анализ значимости коэффициентов эластичности
            if hasattr(self.MR, 'PAC') and self.MR.PAC is not None:
                f.write("АНАЛИЗ ВЛИЯНИЯ ПЕРЕМЕННЫХ:\n")
                f.write("-" * 35 + "\n")
                sorted_pac = sorted(zip(X_ADD, self.MR.PAC), key=lambda x: abs(x[1]), reverse=True)

                f.write("Ранжирование по влиянию на результат:\n")
                for i, (var_name, pac) in enumerate(sorted_pac, 1):
                    influence = "Высокое" if (pac) > 0.3 else "Среднее" if abs(pac) > 0.1 else "Низкое"
                    f.write(f"{i}. {var_name}: {pac*100:6.2f}% ({influence} влияние)\n")
                    f.write(f"   → При изменении {var_name} на 1%, Y изменяется на {pac*100:+.2f}%\n")
                f.write("\n")

    def openReport(self, file_name):
        if file_name:
            try:
                with open(file_name, 'r', encoding='utf-8') as f:
                                content = f.read()

                self.ui.reportText.clear()
                # self.ui.reportText.append(f'Чтение данных из файла: {file_name}\n')
                self.ui.reportText.setFont(QFont('Consolas', 10))
                self.ui.reportText.append(content)

            except Exception as e:
                self.statusBar.showMessage(f"Ошибка открытия файла: {str(e)}")

    ##############################################
    # [ Работа объекта регрессии analyseWindow ] #
    ##############################################

    # Инициализатор объекта регрессии
    def MRInit(self, path):
        self.MR = StepwiseRegressionEngine(path,1)   # Создание объекта регрессии (файл и номер страницы)
        self.MR.readExel(1)                     # Чтение файла (4 - зависимые переменные)
        if self.MR.dataset.size == 0:           # !! - Обработка ошибки - !!
            return print("MRInit")

        self.ui.dockAnalyse.setVisible(True)
        self.ui.dockAnalyse.setStyleSheet(StyleSheetScrollBar)
        self.ui.choiceYComboBox.addItems(self.MR.dataset.columns[self.MR.ColumnsX:self.MR.Columns])     # Список зависимых переменных
        self.ui.countYComboBox.addItems(map(str, range(1, self.MR.Columns - 1)))

        self.actionsDWA()   # Обработка кнопок dockWidget Analyse
        self.actionsRW()   # Обработка кнопок Report Widget
        self.actionsDWG()   # Обработка кнопок dockWidget Graphics
        self.setTableDatabase() # Вывод данных из Exel в таблицу
        self.setCountDepVar() # Выбор количества зависимых переменных
        self.setChoiceDepVar()  # Выбор зависимых переменных
        self.setEditTextBox() # Вывод editBoxes с расчетными параметрами регрессии
        # self.graphicsDraw() # Отрисовка графика

    #------------------------------------------#
    # Изменение цвета колонн таблицы
    def changeBackgrounColor(self,columns_paint,colorBG,colorFG):
        for i in range(self.MR.Lines):
            for j in columns_paint:
                self.ui.tableDatabase.item(i,j).setBackground(QtGui.QColor(colorBG)) # Раскрашивание фона столбцов
                self.ui.tableDatabase.item(i,j).setForeground(QtGui.QColor(colorFG)) # Раскрашивание текста столбцов

    # Редактирование полей с параметрами регрессии
    def setEditTextBox(self):
        try:
            self.ui.lineEditKK.setText(str(self.MR.R))
            self.ui.lineEditKD.setText(str(self.MR.R2))
            self.ui.lineEditKS.setText(str(self.MR.FSKF))

            self.ui.lineEditOD.setText(str(self.MR.Y_Se2))
            self.ui.lineEditNOD.setText(str(self.MR.Y_S2))
            self.ui.lineEditOSKO.setText(str(self.MR.Y_S))

            self.ui.lineEditKDConst.setText(str(self.MR.Const_R2))
            self.ui.lineEditKSConst.setText(str(self.MR.Const_FSKF))

            # Один обработчик для всех lineEdit
            self.ui.lineEditKDConst.textChanged.connect(self.on_float_text_changed)
            self.ui.lineEditKSConst.textChanged.connect(self.on_float_text_changed)

        except Exception as e:
                self.statusBar.showMessage(f"Ошибка записи параметров: {str(e)}")

    def on_float_text_changed(self, text):
        sender = self.sender()  # Получаем объект, который отправил сигнал

        if sender == self.ui.lineEditKDConst:
            try:
                self.MR.Const_R2 = float(text.replace(',', '.'))
                self.statusBar.showMessage(f"Const_R2 обновлен: {self.MR.Const_R2}")
            except ValueError:
                self.statusBar.showMessage("Ошибка: введите корректное число для Const_R2")

        elif sender == self.ui.lineEditKSConst:
            try:
                self.MR.Const_FSKF = float(text.replace(',', '.'))
                self.statusBar.showMessage(f"Const_FSKF обновлен: {self.MR.Const_FSKF}")
            except ValueError:
                self.statusBar.showMessage("Ошибка: введите корректное число для Const_FSKF")

    ##############################################
    # [ Работа c tableDatabase ] #
    ##############################################
    def setTableDatabase(self):
        self.ui.tableDatabase.setStyleSheet(StyleSheetTableDatabase + StyleSheetScrollBar)
        self.ui.tableDatabase.setRowCount(self.MR.Lines)                               # Количество строк
        self.ui.tableDatabase.setColumnCount(self.MR.Columns)                          # Количество столбцов
        self.ui.tableDatabase.setHorizontalHeaderLabels(self.MR.dataset.columns);      # Первая строка таблицы (заголовки столбцов)

        # Вывод таблицы из Exel
        for row in self.MR.dataset.iterrows():
            values = row[1]
            for col_index, value in enumerate(values):
                tableItem = QTableWidgetItem(str(value))
                self.ui.tableDatabase.setItem(row[0],col_index,tableItem)

        self.ui.tableDatabase.itemSelectionChanged.connect(self.tableDatabaseSelectionItem)
        self.ui.buttonAdd.clicked.connect(self.userStepwiseStep) # Включение независимого фактора по мнению пользователя
        self.ui.buttonDelete.clicked.connect(self.userBackwardStep) # Включение независимого фактора по мнению пользователя

    def tableDatabaseSelectionItem(self):
        selected_items = self.ui.tableDatabase.selectedItems()
        if selected_items:
            selected_item = selected_items[0]  # Берем первый выделенный элемент
            column_number = selected_item.column()
            if(column_number in self.MR.IndX_ADD):
                self.ui.buttonDelete.setEnabled(True)
                self.ui.buttonAdd.setEnabled(False)
                self.ui.buttonDelete.setStyleSheet("""#buttonDelete{border-color: #ffc5c5;} #buttonDelete:hover{border-color: #505A6E;}""")
                self.ui.buttonAdd.setStyleSheet("""#buttonAdd{border-color: white;}""")
            elif(column_number in self.MR.IndX_DEL):
                self.ui.buttonAdd.setEnabled(True)
                self.ui.buttonDelete.setEnabled(False)
                self.ui.buttonDelete.setStyleSheet("""#buttonDelete{border-color: white;}""")
                self.ui.buttonAdd.setStyleSheet("""#buttonAdd{border-color: #bcf5d3;} #buttonAdd:hover{border-color: #505A6E;}""")
            else:
                self.ui.buttonAdd.setEnabled(False)
                self.ui.buttonDelete.setEnabled(False)
                self.ui.buttonDelete.setStyleSheet("""#buttonDelete{border-color: white;}""")
                self.ui.buttonAdd.setStyleSheet("""#buttonAdd{border-color: white;}""")
            return column_number



if __name__ == '__main__':
    import sys
    from PyQt5.QtWidgets import QApplication
    app = QApplication(sys.argv)
    app.setStyleSheet(StyleSheetApp)
    w = MainWindow()
    w.setWindowIcon(QIcon(r'src\ui\resources\Logo\SHV_icon.svg'))
    w.show()
    # w = FramelessWindow()
    # w.setWindowTitle(' Множественная регрессия [by Artem Shishov]')
    # w.setIconSize(40)
    # w.setWindowIcon(QIcon(r'src\ui\resources\Logo\SHV_icon.svg'))
    # w.setWidget(MainWindow(w))          # Добавить свое окно
    # w.show()
    sys.exit(app.exec_())